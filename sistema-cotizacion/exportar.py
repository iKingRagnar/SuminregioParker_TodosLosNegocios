# -*- coding: utf-8 -*-
"""Exportación a Excel y PDF - sin dependencias de pago."""
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    REPORTLAB = True
except ImportError:
    REPORTLAB = False

import config


def exportar_cotizacion_excel(cotizacion, lineas, cliente, ruta_archivo):
    """Genera archivo .xlsx con la cotización."""
    if not OPENPYXL:
        raise ImportError("Instala openpyxl: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Membrete
    ws.merge_cells("A1:F1")
    ws["A1"] = config.EMPRESA_NOMBRE
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:F2")
    ws["A2"] = config.EMPRESA_DIRECCION or ""
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Cotización: {cotizacion.get('folio', '')}"

    fila = 5
    ws[f"A{fila}"] = "Cliente:"
    ws[f"A{fila}"].font = Font(bold=True)
    ws[f"B{fila}"] = cliente.get("nombre", "")
    fila += 1
    ws[f"A{fila}"] = "RFC:"
    ws[f"B{fila}"] = cliente.get("rfc", "")
    fila += 1
    ws[f"A{fila}"] = "Dirección:"
    ws[f"B{fila}"] = cliente.get("direccion", "")
    fila += 2

    # Encabezados tabla
    encabezados = ["Código", "Descripción", "Cantidad", "Precio Unit.", "Subtotal", "Total"]
    for c, enc in enumerate(encabezados, 1):
        ws.cell(row=fila, column=c, value=enc)
        ws.cell(row=fila, column=c).font = Font(bold=True)
    fila += 1

    for lin in lineas:
        ws.cell(row=fila, column=1, value=lin.get("codigo") or "")
        ws.cell(row=fila, column=2, value=lin.get("descripcion", ""))
        ws.cell(row=fila, column=3, value=lin.get("cantidad", 0))
        ws.cell(row=fila, column=4, value=lin.get("precio_unitario", 0))
        ws.cell(row=fila, column=5, value=lin.get("subtotal", 0))
        ws.cell(row=fila, column=6, value=lin.get("total", 0))
        fila += 1

    fila += 1
    ws.cell(row=fila, column=5, value="Subtotal:")
    ws.cell(row=fila, column=5).font = Font(bold=True)
    ws.cell(row=fila, column=6, value=cotizacion.get("subtotal", 0))
    fila += 1
    ws.cell(row=fila, column=5, value="IVA:")
    ws.cell(row=fila, column=5).font = Font(bold=True)
    ws.cell(row=fila, column=6, value=cotizacion.get("iva", 0))
    fila += 1
    ws.cell(row=fila, column=5, value="Total:")
    ws.cell(row=fila, column=5).font = Font(bold=True)
    ws.cell(row=fila, column=6, value=cotizacion.get("total", 0))

    wb.save(ruta_archivo)
    return ruta_archivo


def exportar_cotizacion_pdf(cotizacion, lineas, cliente, ruta_archivo):
    """Genera PDF con la cotización."""
    if not REPORTLAB:
        raise ImportError("Instala reportlab: pip install reportlab")
    doc = SimpleDocTemplate(
        ruta_archivo,
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(config.EMPRESA_NOMBRE, styles["Title"]))
    story.append(Paragraph(config.EMPRESA_DIRECCION or "", styles["Normal"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Cotización: {cotizacion.get('folio', '')}", styles["Heading2"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Cliente:</b> {cliente.get('nombre', '')}", styles["Normal"]))
    story.append(Paragraph(f"<b>RFC:</b> {cliente.get('rfc', '')}", styles["Normal"]))
    story.append(Paragraph(f"<b>Dirección:</b> {cliente.get('direccion', '')}", styles["Normal"]))
    story.append(Spacer(1, 20))

    data = [["Código", "Descripción", "Cant.", "P.Unit.", "Subtotal", "Total"]]
    for lin in lineas:
        data.append([
            lin.get("codigo") or "",
            (lin.get("descripcion") or "")[:40],
            str(lin.get("cantidad", 0)),
            f"${lin.get('precio_unitario', 0):,.2f}",
            f"${lin.get('subtotal', 0):,.2f}",
            f"${lin.get('total', 0):,.2f}",
        ])
    data.append(["", "", "", "", "Subtotal:", f"${cotizacion.get('subtotal', 0):,.2f}"])
    data.append(["", "", "", "", "IVA:", f"${cotizacion.get('iva', 0):,.2f}"])
    data.append(["", "", "", "", "Total:", f"${cotizacion.get('total', 0):,.2f}"])

    t = Table(data, colWidths=[1.2*inch, 2.5*inch, 0.5*inch, 0.8*inch, 1*inch, 1*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -4), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (4, -3), (-1, -1), "RIGHT"),
    ]))
    story.append(t)
    doc.build(story)
    return ruta_archivo


def exportar_mantenimientos_excel(maquina_nombre, registros, ruta_archivo):
    """Exporta historial de mantenimientos de una máquina a Excel."""
    if not OPENPYXL:
        raise ImportError("Instala openpyxl: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"
    ws["A1"] = f"Mantenimientos - {maquina_nombre}"
    ws["A1"].font = Font(bold=True, size=12)
    encabezados = ["Fecha", "Tipo", "Descripción", "Técnico", "Horas", "Costo"]
    for c, enc in enumerate(encabezados, 1):
        ws.cell(row=3, column=c, value=enc)
        ws.cell(row=3, column=c).font = Font(bold=True)
    for i, r in enumerate(registros, 4):
        ws.cell(row=i, column=1, value=r.get("fecha_fin") or r.get("fecha_inicio") or "")
        ws.cell(row=i, column=2, value=r.get("tipo", ""))
        ws.cell(row=i, column=3, value=(r.get("descripcion_falla") or r.get("accion_tomada") or "")[:50])
        ws.cell(row=i, column=4, value=r.get("tecnico", ""))
        ws.cell(row=i, column=5, value=r.get("horas_invertidas", 0))
        ws.cell(row=i, column=6, value=r.get("costo_total", 0))
    wb.save(ruta_archivo)
    return ruta_archivo
